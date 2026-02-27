import streamlit as st
import pandas as pd

from openpyxl import Workbook, load_workbook
import os
from datetime import datetime


# 初回アクセス時に開始時刻を保存
if "start_time" not in st.session_state:
    st.session_state.start_time = datetime.now()


# タイトル
st.title("簡単アンケートフォーム")


# 入力項目
name = st.text_input("お名前を入力してください")

age = st.number_input("年齢を入力してください", min_value=0, max_value=120)

satisfaction = st.radio(
    "満足度を教えてください",
    ("とても満足", "満足", "普通", "不満", "とても不満")
)


# 送信ボタン
if st.button("送信"):

    # 終了時刻
    end_time = datetime.now()

    # 開始時刻
    start_time = st.session_state.start_time

    # 所要時間（秒）
    duration = (end_time - start_time).total_seconds()

    # ファイル名
    file_name = "results.xlsx"


    # ファイルがなければ作成
    if not os.path.exists(file_name):

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "回答データ"

        # ヘッダー追加
        sheet.append([
            "名前",
            "年齢",
            "満足度",
            "回答開始時刻",
            "回答終了時刻",
            "回答時間（秒）"
        ])

        workbook.save(file_name)


    # ファイルを開く
    workbook = load_workbook(file_name)
    sheet = workbook.active


    # データ追加
    sheet.append([
        name,
        age,
        satisfaction,
        start_time.strftime("%Y-%m-%d %H:%M:%S"),
        end_time.strftime("%Y-%m-%d %H:%M:%S"),
        round(duration, 2)
    ])


    workbook.save(file_name)


    st.success("回答ありがとうございました！")

    import pandas as pd
import os

FILE_NAME = "results.xlsx"

COLUMNS = [
    "送信日時",
    "名前",
    "年齢",
    "満足度",
    "開始時刻",
    "終了時刻",
    "回答時間(秒)"
]

if not os.path.exists(FILE_NAME):
    df = pd.DataFrame(columns=COLUMNS)
    df.to_excel(FILE_NAME, index=False)
